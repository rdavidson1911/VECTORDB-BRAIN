#!/usr/bin/env python3
"""Generate internal Docker/WSL2 resource budget workbook for OmniKB.

Requires: pip install openpyxl

Output: internal_docs/docker-resource-budget-model.xlsx
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = REPO_ROOT / "internal_docs" / "docker-resource-budget-model.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="E2E8F0", bold=True)
TITLE_FONT = Font(bold=True, size=12)


def _style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        length = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                length = max(length, len(str(val)))
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def sheet_guide(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Guide"
    lines = [
        ("OmniKB — Docker Desktop / WSL2 resource model", ""),
        ("Generated", datetime.now(UTC).isoformat()),
        ("", ""),
        (
            "Purpose",
            "Compare EXPECTED vs ACTUAL CPU/RAM for omnikb-api and omnikb-qdrant on Win11 + WSL2.",
        ),
        ("", ""),
        ("How to use", ""),
        ("1", "Fill DockerDesktop and WSL2 with your machine limits."),
        ("2", "Review OmniKB_Expected (edit bands if you learn better numbers)."),
        ("3", "During idle/ingest/query, run: docker stats --no-stream"),
        ("4", "Log rows in Measurements (one row per sample)."),
        ("5", "Compare sheet shows variance vs expected midpoints."),
        ("", ""),
        ("Doc", "docs/internal/docker-desktop-wsl2-resources.md"),
        ("Regenerate", "python scripts/generate_docker_resource_model.py"),
    ]
    for r, (a, b) in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
    ws.cell(row=1, column=1).font = TITLE_FONT
    _autosize(ws)


def sheet_docker_desktop(wb: Workbook) -> None:
    ws = wb.create_sheet("DockerDesktop")
    headers = ["Setting", "Value", "Unit", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    rows = [
        ("Host OS", "Windows 11 Pro", "", "Team standard"),
        ("Docker Desktop CPUs", 8, "count", "Settings → Resources"),
        ("Docker Desktop Memory", 16, "GB", "VM cap for all containers"),
        ("Docker Desktop Swap", 4, "GB", "If enabled"),
        ("Disk image size", 64, "GB", "Settings → Resources → Disk"),
        ("Compose project", "VECTORDB-BRAIN", "", "docker compose up"),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _autosize(ws)


def sheet_wsl2(wb: Workbook) -> None:
    ws = wb.create_sheet("WSL2")
    headers = ["Setting", "Value", "Unit", "Notes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    rows = [
        ("wsl2.memory", "16GB", "string", "%UserProfile%\\.wslconfig"),
        ("wsl2.processors", 8, "count", "Should align with Docker CPUs"),
        ("wsl2.swap", "8GB", "string", "Optional"),
        ("Default distro", "docker-desktop", "", "wsl -l -v"),
        ("After config change", "wsl --shutdown", "", "Required"),
    ]
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    _autosize(ws)


def sheet_expected(wb: Workbook) -> None:
    ws = wb.create_sheet("OmniKB_Expected")
    headers = [
        "Service",
        "Scenario",
        "CPU_%_min",
        "CPU_%_typical",
        "CPU_%_max",
        "RAM_MB_min",
        "RAM_MB_typical",
        "RAM_MB_max",
        "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    # Heuristic bands for planning — replace with your measured baselines.
    data = [
        ("omnikb-qdrant", "idle", 0.5, 2, 8, 256, 512, 1200, "Small corpus on SSD"),
        ("omnikb-qdrant", "ingest", 5, 25, 60, 400, 900, 2500, "Upsert + indexing"),
        ("omnikb-qdrant", "query", 2, 15, 40, 300, 700, 2000, "Search load"),
        ("omnikb-api", "idle", 0.2, 1, 5, 400, 800, 1500, "Model may stay resident"),
        ("omnikb-api", "ingest_cold", 20, 80, 200, 1500, 3500, 8000, "First embed + torch"),
        ("omnikb-api", "ingest_warm", 10, 50, 120, 800, 2000, 5000, "After model loaded"),
        ("omnikb-api", "query", 5, 30, 100, 600, 1500, 4000, "Encode + API overhead"),
        ("vmmem (WSL VM)", "ingest_cold", 15, 50, 95, 8000, 12000, 24000, "Host-level vmmem WS"),
    ]
    for r, row in enumerate(data, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        ws.cell(row=r, column=10, value=f'=A{r}&"|"&B{r}')
    ws.cell(row=1, column=10, value="lookup_key")
    _autosize(ws)


def sheet_measurements(wb: Workbook) -> None:
    ws = wb.create_sheet("Measurements")
    headers = [
        "timestamp_utc",
        "scenario",
        "duration_min",
        "api_cpu_pct",
        "qdrant_cpu_pct",
        "api_mem_mb",
        "qdrant_mem_mb",
        "compose_mem_mb_sum",
        "vmmem_ws_mb",
        "docker_mem_limit_gb",
        "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    # Example row (delete or overwrite)
    example = [
        datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "idle",
        2,
        0.5,
        0.3,
        850,
        420,
        1270,
        14500,
        16,
        "docker stats --no-stream",
    ]
    for c, val in enumerate(example, start=1):
        ws.cell(row=2, column=c, value=val)
    _autosize(ws)


def sheet_compare(wb: Workbook) -> None:
    ws = wb.create_sheet("Compare")
    ws.cell(
        row=1,
        column=1,
        value="Link Measurements to OmniKB_Expected (lookup_key = service|scenario)",
    ).font = TITLE_FONT
    headers = [
        "meas_row",
        "scenario",
        "service",
        "actual_ram_mb",
        "expected_ram_typical",
        "ram_delta_mb",
        "actual_cpu_pct",
        "expected_cpu_typical",
        "cpu_delta_pct",
        "within_ram_band",
        "notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    _style_header_row(ws, 3, len(headers))

    def _write_compare_row(r: int, meas_row: int, service: str, ram_col: str, cpu_col: str) -> None:
        key = f'$C{r}&"|"&$B{r}'
        ws.cell(row=r, column=1, value=meas_row)
        ws.cell(row=r, column=2, value=f"=Measurements!B{meas_row}")
        ws.cell(row=r, column=3, value=service)
        ws.cell(row=r, column=4, value=f"=Measurements!{ram_col}{meas_row}")
        ws.cell(
            row=r,
            column=5,
            value=f'=IFERROR(INDEX(OmniKB_Expected!$G:$G,MATCH({key},OmniKB_Expected!$J:$J,0)),"")',
        )
        ws.cell(row=r, column=6, value=f'=IF(E{r}="","",D{r}-E{r})')
        ws.cell(row=r, column=7, value=f"=Measurements!{cpu_col}{meas_row}")
        ws.cell(
            row=r,
            column=8,
            value=f'=IFERROR(INDEX(OmniKB_Expected!$D:$D,MATCH({key},OmniKB_Expected!$J:$J,0)),"")',
        )
        ws.cell(row=r, column=9, value=f'=IF(H{r}="","",G{r}-H{r})')
        ws.cell(
            row=r,
            column=10,
            value=(
                f'=IF(E{r}="","",IF(AND(D{r}>=INDEX(OmniKB_Expected!$F:$F,MATCH({key},OmniKB_Expected!$J:$J,0)),'
                f'D{r}<=INDEX(OmniKB_Expected!$H:$H,MATCH({key},OmniKB_Expected!$J:$J,0))),"yes","no"))'
            ),
        )

    _write_compare_row(4, 2, "omnikb-api", "F", "D")
    _write_compare_row(5, 2, "omnikb-qdrant", "G", "E")
    ws.cell(
        row=4,
        column=11,
        value="Copy rows 4–5 down for each Measurements data row; update meas_row.",
    )
    _autosize(ws)


def main() -> None:
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet_guide(wb)
    sheet_docker_desktop(wb)
    sheet_wsl2(wb)
    sheet_expected(wb)
    sheet_measurements(wb)
    sheet_compare(wb)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
